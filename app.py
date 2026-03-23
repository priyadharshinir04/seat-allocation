import os
import csv
from datetime import datetime, timedelta
from flask import Flask, render_template, request, redirect, url_for, flash, jsonify, session, send_file
from werkzeug.utils import secure_filename
from itertools import combinations
import io
import qrcode
from PIL import Image
import logging
from logging.handlers import RotatingFileHandler
from twilio.rest import Client
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

app = Flask(__name__)
app.secret_key = "college_seating_secret_key"
UPLOAD_FOLDER = 'uploads'
ALLOWED_EXTENSIONS = {'csv', 'xlsx', 'xls'}

app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

# Configure session to persist data across page reloads and navigation
app.config['PERMANENT_SESSION_LIFETIME'] = timedelta(hours=24)
app.config['SESSION_REFRESH_EACH_REQUEST'] = True

# Setup logging for SMS notifications
# Use /tmp for Render (ephemeral storage) or local logs for local dev
import tempfile
log_dir = os.path.join(tempfile.gettempdir(), 'flask_logs')
if not os.path.exists(log_dir):
    os.makedirs(log_dir)

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        RotatingFileHandler(os.path.join(log_dir, 'sms_notifications.log'), maxBytes=10485760, backupCount=10),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

# Twilio Configuration (Use environment variables in production)
TWILIO_ACCOUNT_SID = os.environ.get('TWILIO_ACCOUNT_SID', 'YOUR_ACCOUNT_SID_HERE')
TWILIO_AUTH_TOKEN = os.environ.get('TWILIO_AUTH_TOKEN', 'YOUR_AUTH_TOKEN_HERE')
TWILIO_PHONE_NUMBER = os.environ.get('TWILIO_PHONE_NUMBER', '+1234567890')

# SMS Demo Mode - Set to True for demo/presentation usage
# When True: Simulates SMS sending without calling Twilio API (perfect for presentations and trials)
# Set to False only when Twilio SMS is fully configured and authorized for your region
SMS_DEMO_MODE = os.environ.get('SMS_DEMO_MODE', 'True').lower() == 'true'

# VERIFIED RECIPIENT ALLOWLIST for Twilio Trial Account
# Trial accounts can only send SMS to verified phone numbers
# Add verified/test phone numbers here in E.164 format (+91XXXXXXXXXX)
VERIFIED_TEST_NUMBERS = [
    "+919360792531",  # User's verified test number
]

# Ensure upload folder exists
if not os.path.exists(UPLOAD_FOLDER):
    os.makedirs(UPLOAD_FOLDER)

# In-memory storage for simplicity (replace with DB for production)
students_data = []
allocation_results = []
exam_schedules = []  # New: Store exam schedule data

# Session storage for on-campus configuration
exam_config = {}

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def get_exam_schedule_for_student(student_year, student_dept):
    """Get exam schedule for a specific student based on year and department"""
    global exam_schedules
    exam_schedules_from_session = session.get('exam_schedules', [])
    schedules = exam_schedules_from_session if exam_schedules_from_session else exam_schedules
    
    if not schedules:
        return None
    
    # Match by year and department
    for schedule in schedules:
        if str(schedule.get('year', '')) == str(student_year):
            dept = schedule.get('department', 'ALL')
            # Match if department is ALL (common for year 1) or matches exactly
            if dept == 'ALL' or dept == student_dept:
                return schedule
    
    return None

def validate_phone_number(phone):
    """
    Validate and normalize phone number to E.164 format.
    Returns: (is_valid, normalized_phone, error_description)
    
    Examples:
    - "9360792531" → (True, "+919360792531", None)
    - "+91 9360792531" → (True, "+919360792531", None)
    - "919360792531" → (True, "+919360792531", None)
    - "" → (False, None, "Empty phone number")
    - "123" → (False, None, "Invalid format (too short)")
    """
    if not phone:
        return False, None, "Empty phone number"
    
    phone = str(phone).strip()
    
    # Step 1: Remove common separators (spaces, dashes, parentheses)
    phone = phone.replace(' ', '').replace('-', '').replace('(', '').replace(')', '')
    
    # Step 2: Extract digits only
    digits_only = ''.join(c for c in phone if c.isdigit())
    
    # Step 3: Check minimum length (10 digits for Indian numbers)
    if len(digits_only) < 10:
        return False, None, f"Invalid format - too short (needs 10+ digits, got {len(digits_only)})"
    
    # Step 4: Normalize to E.164 format
    if phone.startswith('+'):
        # Already has +, keep as is but validate
        normalized = phone
    elif phone.startswith('91') and len(digits_only) == 12:
        # Format: 919360792531 → +919360792531
        normalized = '+' + digits_only
    elif phone.startswith('0') and len(digits_only) == 11:
        # Format: 09360792531 → +919360792531 (Indian leading 0)
        normalized = '+91' + digits_only[1:]
    elif len(digits_only) == 10:
        # Format: 9360792531 → +919360792531 (Indian 10-digit)
        normalized = '+91' + digits_only
    else:
        # Unknown format with country code
        normalized = '+' + digits_only
    
    # Step 5: Validate E.164 format: +<country_code><number>
    # For Indian numbers: +91XXXXXXXXXX (12 chars total, 10 digits after +91)
    if not normalized.startswith('+'):
        return False, None, "Invalid format - must start with +91"
    
    normalized_digits = normalized[1:]
    if len(normalized_digits) < 10:
        return False, None, f"Invalid format - too short after normalization"
    
    return True, normalized, None

def send_sms_notification(phone_number, student_name, subject_name, exam_time, room_number, bench_number, index=0):
    """
    Send SMS notification to student before exam.
    
    Returns: (success_bool, message, status_type, skip_reason)
    Status types: 'success', 'failed', 'skipped'
    
    For Twilio trial accounts:
    - Only sends to verified numbers in VERIFIED_TEST_NUMBERS
    - Skips all other numbers gracefully (not counted as failures)
    - Logs detailed reason for each skip/failure
    """
    try:
        # STEP 1: Validate and normalize phone number
        is_valid, normalized_phone, validation_error = validate_phone_number(phone_number)
        
        if not is_valid:
            logger.warning(f"📵 VALIDATION FAILED - {student_name}: {validation_error}")
            return False, validation_error, "skipped", "Invalid phone format"
        
        # STEP 2: Check if Twilio credentials are configured
        if TWILIO_ACCOUNT_SID == 'YOUR_ACCOUNT_SID_HERE' or not TWILIO_AUTH_TOKEN:
            logger.warning(f"📵 CREDENTIALS MISSING - {student_name} ({normalized_phone}): Skipped")
            return False, "Twilio credentials not configured", "skipped", "Config missing"
        
        # STEP 3: CRITICAL - Check if number is verified (trial account restriction)
        # This is the key performance optimization - skip API call if not verified
        if normalized_phone not in VERIFIED_TEST_NUMBERS:
            reason = "Trial restriction - unverified recipient (add to VERIFIED_TEST_NUMBERS to send)"
            logger.info(f"⏭️  SKIPPED (TRIAL RESTRICTION) - {student_name} ({normalized_phone})")
            return False, reason, "skipped", reason
        
        # STEP 4: Build SMS message (only after all validations pass)
        message_body = f"""Hi {student_name},
Reminder: Your {subject_name} exam starts at {exam_time}.
Room: {room_number}, Bench: {bench_number}.
Please arrive 45 minutes early."""
        
        logger.info(f"📤 ATTEMPTING SMS - {student_name} ({normalized_phone})...")
        
        # STEP 5: In DEMO mode, simulate SMS (don't call Twilio API)
        if SMS_DEMO_MODE:
            logger.info(f"✅ DEMO SIMULATION - SMS would be sent to {student_name} ({normalized_phone})")
            logger.info(f"   Message: {message_body[:50]}...")
            return True, f"Simulated", "success", None
        
        # STEP 5: Call Twilio API (only for verified numbers)
        try:
            client = Client(TWILIO_ACCOUNT_SID, TWILIO_AUTH_TOKEN)
            message = client.messages.create(
                body=message_body,
                from_=TWILIO_PHONE_NUMBER,
                to=normalized_phone
            )
            
            logger.info(f"✅ SMS SENT SUCCESS - {student_name} ({normalized_phone}) | SID: {message.sid}")
            return True, f"Sent", "success", None
        
        except Exception as api_error:
            error_msg = str(api_error)
            
            # Parse Twilio error codes to provide user-friendly messages
            if '21211' in error_msg:
                error_msg = "Invalid phone number (Twilio rejected format)"
            elif '21612' in error_msg or '21614' in error_msg or '21617' in error_msg:
                error_msg = "Invalid phone number format"
            elif '21610' in error_msg:
                error_msg = "Account not authorized - check Twilio account status"
            elif '63018' in error_msg:
                error_msg = "Account suspended"
            elif '63038' in error_msg or '63000' in error_msg:
                error_msg = "SMS permission denied (Account not ready to send SMS - check Twilio Console > SMS > Geographical Permissions)"
            elif '429' in error_msg:
                error_msg = "Rate limited or permission error - check Twilio account SMS capability"
            elif 'Unauthorized' in error_msg or '401' in error_msg:
                error_msg = "Twilio authentication failed - verify Account SID & Auth Token"
            
            logger.error(f"❌ TWILIO API ERROR - {student_name} ({normalized_phone}): {error_msg} [Raw: {str(api_error)[:100]}]")
            return False, error_msg, "failed", error_msg
    
    except Exception as e:
        error_msg = str(e)
        logger.error(f"❌ UNEXPECTED ERROR - {student_name} ({phone_number}): {error_msg}")
        return False, error_msg, "failed", error_msg

def send_bulk_sms_notifications():
    """
    Send SMS notifications to all students with upcoming exams.
    
    Implements:
    - Phone number validation and normalization to E.164 format
    - Verified recipient allowlist (skips unverified in trial accounts)
    - Detailed error logging per student
    - Performance optimization (skips unverified before API call)
    - Clean separation of failed vs skipped counts
    """
    try:
        # Get allocation results from session or global
        alloc_results = session.get('allocation_results') or allocation_results
        
        if not alloc_results:
            logger.warning("⚠️  No students allocated for SMS sending")
            return {
                "status": "error",
                "success_count": 0,
                "failed_count": 0,
                "skipped_count": 0,
                "total": 0,
                "message": "No students allocated yet",
                "verified_count": len(VERIFIED_TEST_NUMBERS)
            }
        
        config = session.get('exam_config', {})
        exam_mode = config.get('exam_type', 'Internal')
        exam_time = config.get('exam_time', 'Not Set')
        
        logger.info("\n" + "="*70)
        logger.info("🚀 SMS NOTIFICATION CAMPAIGN STARTED - SAFE DEMO MODE")
        logger.info("="*70)
        logger.info(f"📅 Exam Mode: {exam_mode}")
        logger.info(f"⏰ Exam Time: {exam_time}")
        logger.info(f"👥 Total Students: {len(alloc_results)}")
        logger.info(f"✅ Verified Recipients: {len(VERIFIED_TEST_NUMBERS)}")
        logger.info(f"🔐 Verified Numbers: {', '.join(VERIFIED_TEST_NUMBERS)}")
        if SMS_DEMO_MODE:
            logger.info(f"🎯 DEMO MODE: ON - Professional Demo (Simulating SMS delivery)")
        else:
            logger.info(f"🎯 PRODUCTION MODE: Real Twilio SMS API enabled")
        logger.info("="*70 + "\n")
        
        success_count = 0
        failed_count = 0
        skipped_count = 0
        errors = []
        skipped_details = []
        
        for idx, student in enumerate(alloc_results):
            phone_number = student.get('phone_number', '').strip()
            student_name = student.get('candidate_name', 'Student')
            subject_name = student.get('subject_name', 'Unknown Subject')
            room_number = student.get('room_number', 'N/A')
            bench_number = student.get('bench_number', 'N/A')
            
            # Send SMS (handles validation, verified check, and Twilio API)
            success, message, status, skip_reason = send_sms_notification(
                phone_number,
                student_name,
                subject_name,
                exam_time,
                room_number,
                bench_number,
                index=idx
            )
            
            if status == "success":
                success_count += 1
            elif status == "failed":
                failed_count += 1
                errors.append({
                    "student": student_name,
                    "phone": phone_number,
                    "error": message
                })
            else:  # skipped
                skipped_count += 1
                skipped_details.append({
                    "student": student_name,
                    "phone": phone_number,
                    "reason": skip_reason if skip_reason else message
                })
        
        # Build comprehensive result
        # Build comprehensive result
        if SMS_DEMO_MODE:
            # In demo mode, show clean result (all successful simulations)
            result = {
                "status": "success",
                "success_count": success_count,
                "failed_count": failed_count,
                "skipped_count": skipped_count,
                "total": len(allocation_results),
                "message": f"✅ {success_count} simulated | ❌ {failed_count} failed | ⏭️  {skipped_count} skipped",
                "errors": errors[:5],
                "skipped_details": skipped_details[:10],
                "demo_mode": True,
                "demo_message": "Demo Mode: SMS delivery simulated successfully"
            }
        else:
            # Production mode
            result = {
                "status": "success" if failed_count == 0 else ("partial" if success_count > 0 else "error"),
                "success_count": success_count,
                "failed_count": failed_count,
                "skipped_count": skipped_count,
                "total": len(alloc_results),
                "message": f"✅ {success_count} sent | ❌ {failed_count} failed | ⏭️  {skipped_count} skipped",
                "errors": errors[:5],
                "skipped_details": skipped_details[:10],
                "demo_mode": False
            }
        
        # Log final summary
        if SMS_DEMO_MODE:
            logger.info("📊 SMS CAMPAIGN COMPLETE - DEMO MODE")
        else:
            logger.info("📊 SMS CAMPAIGN COMPLETE - PRODUCTION MODE")
        logger.info("="*70)
        logger.info(f"✅ Successfully Simulated/Sent:  {success_count}")
        logger.info(f"❌ Failed:                       {failed_count}")
        logger.info(f"⏭️  Skipped Total:               {skipped_count}")
        logger.info(f"📋 Total Processed:              {len(alloc_results)}")
        logger.info("="*70 + "\n")
        
        return result
    
    except Exception as e:
        logger.error(f"\n❌ FATAL ERROR: {str(e)}\n")
        return {
            "status": "error",
            "success_count": 0,
            "failed_count": 0,
            "skipped_count": 0,
            "total": 0,
            "message": f"Operation failed: {str(e)}",
            "verified_count": len(VERIFIED_TEST_NUMBERS)
        }

def read_csv_to_dict(filepath):
    """Read CSV file into list of dicts"""
    try:
        data = []
        with open(filepath, 'r', encoding='utf-8') as f:
            reader = csv.DictReader(f)
            for row in reader:
                data.append(row)
        return data
    except Exception as e:
        return None

def read_excel_to_dict(filepath):
    """Read Excel file into list of dicts"""
    try:
        wb = load_workbook(filepath)
        ws = wb.active
        
        # Get headers from first row
        headers = []
        for cell in ws[1]:
            headers.append(cell.value)
        
        # Read data rows
        data = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            row_dict = {}
            for i, header in enumerate(headers):
                row_dict[header] = row[i] if i < len(row) else None
            data.append(row_dict)
        
        return data
    except Exception as e:
        return None

def validate_student_data(data):
    """Validate and clean student data - works with list of dicts"""
    if not data or not isinstance(data, list):
        return None, "Invalid data format"
    
    # Normalize column names
    normalized_data = []
    column_map = {}
    
    if data:
        first_row = data[0]
        for col in first_row.keys():
            col_lower = col.lower().strip()
            if col_lower == 'register number':
                column_map[col] = 'Register Number'
            elif col_lower in ['candidate name', 'student name', 'name']:
                column_map[col] = 'Candidate Name'
            elif col_lower == 'department':
                column_map[col] = 'Department'
            elif 'year' in col_lower or col_lower in ['batch', 'sem']:
                column_map[col] = 'Year'
            elif 'phone' in col_lower or col_lower in ['mobile', 'contact', 'phone number']:
                column_map[col] = 'Phone Number'
            else:
                column_map[col] = col
    
    # Create normalized rows
    for row in data:
        normalized_row = {}
        for old_col, val in row.items():
            new_col = column_map.get(old_col, old_col)
            normalized_row[new_col] = val
        normalized_data.append(normalized_row)
    
    # Check for required columns
    required_columns = ['Register Number', 'Candidate Name', 'Department']
    if normalized_data:
        missing_cols = [col for col in required_columns if col not in normalized_data[0].keys()]
        if missing_cols:
            return None, f"Missing columns: {', '.join(missing_cols)}"
    
    # Remove rows with missing values in required columns
    cleaned_data = []
    for row in normalized_data:
        if (row.get('Register Number') is not None and row.get('Register Number') != '' and
            row.get('Candidate Name') is not None and row.get('Candidate Name') != '' and
            row.get('Department') is not None and row.get('Department') != ''):
            cleaned_data.append(row)
    
    # Parse Year column if it exists
    if cleaned_data and 'Year' in cleaned_data[0]:
        for row in cleaned_data:
            year_val = row.get('Year')
            if year_val is None or year_val == '':
                row['Year'] = '0'
            else:
                year_str = str(year_val).strip().lower()
                # Handle "2nd Year", "3rd Year", "4th Year", "1st Year" format
                year_str = year_str.replace('st ', '') \
                                  .replace('nd ', '') \
                                  .replace('rd ', '') \
                                  .replace('th ', '') \
                                  .replace('year', '') \
                                  .strip()
                # Extract first digit
                found = False
                for char in year_str:
                    if char.isdigit():
                        row['Year'] = char
                        found = True
                        break
                if not found:
                    row['Year'] = '0'
    
    # Clean phone numbers if present (optional field)
    if cleaned_data and 'Phone Number' in cleaned_data[0]:
        for row in cleaned_data:
            phone_val = row.get('Phone Number')
            if phone_val is None or phone_val == '':
                row['Phone Number'] = ''
            else:
                phone_str = str(phone_val).strip()
                # Remove common separators
                phone_str = phone_str.replace('-', '').replace(' ', '').replace('(', '').replace(')', '')
                row['Phone Number'] = phone_str
    
    # Remove duplicates
    seen = set()
    unique_data = []
    for row in cleaned_data:
        reg_no = row.get('Register Number')
        if reg_no not in seen:
            seen.add(reg_no)
            unique_data.append(row)
    
    return unique_data, None

def internal_exam_allocation(students_list):
    """
    Internal Exam Allocation - 2 students per bench
    - 20 benches per classroom = 40 students per classroom
    - Each bench: 2 students from different departments and different years
    - Sorted by Register Number within each Year group for structured seating
    """
    if not students_list:
        return None, "No students to allocate"
    
    # Convert to list of dicts if needed and sort
    students_data = students_list if isinstance(students_list, list) else list(students_list)
    
    # Sort by Year and Register Number
    students_data = sorted(students_data, 
                          key=lambda x: (str(x.get('Year', '0')), str(x.get('Register Number', ''))))
    
    # Group by Department + Year
    groups = {}
    for student in students_data:
        dept = student.get('Department', '')
        year = str(student.get('Year', '0'))
        key = (dept, year)
        
        if key not in groups:
            groups[key] = []
        
        groups[key].append({
            'register_number': student.get('Register Number', ''),
            'candidate_name': student.get('Candidate Name', ''),
            'department': student.get('Department', ''),
            'year': year,
            'phone_number': student.get('Phone Number', '')
        })
    
    # Create valid pairs (dept1 ≠ dept2 AND year1 ≠ year2)
    valid_pairs = []
    for g1, g2 in combinations(groups.keys(), 2):
        dept1, year1 = g1
        dept2, year2 = g2
        if dept1 != dept2 and year1 != year2:
            valid_pairs.append((g1, g2))
    
    allocation = []
    room_no = 1
    bench_no = 1
    
    # Allocate benches using valid pairs
    for (g1, g2) in valid_pairs:
        if not groups[g1] or not groups[g2]:
            continue
        
        while groups[g1] and groups[g2]:
            # Assign 2 students per bench
            student1 = groups[g1].pop(0)
            student2 = groups[g2].pop(0)
            
            # Left seat
            allocation.append({
                'register_number': student1['register_number'],
                'candidate_name': student1['candidate_name'],
                'department': student1['department'],
                'year': student1['year'],
                'phone_number': student1.get('phone_number', ''),
                'room_number': room_no,
                'bench_number': bench_no,
                'seat_position': 'Left'
            })
            
            # Right seat
            allocation.append({
                'register_number': student2['register_number'],
                'candidate_name': student2['candidate_name'],
                'department': student2['department'],
                'year': student2['year'],
                'phone_number': student2.get('phone_number', ''),
                'room_number': room_no,
                'bench_number': bench_no,
                'seat_position': 'Right'
            })
            
            bench_no += 1
            if bench_no > 20:  # Move to next classroom after 20 benches
                room_no += 1
                bench_no = 1
    
    # Handle remaining students (same-year pairs)
    remaining_groups = {}
    for group_key, group_list in groups.items():
        if group_list:
            remaining_groups[group_key] = group_list
    
    year_based_pairs = []
    for g1, g2 in combinations(remaining_groups.keys(), 2):
        dept1, year1 = g1
        dept2, year2 = g2
        if year1 != year2 and remaining_groups[g1] and remaining_groups[g2]:
            year_based_pairs.append((g1, g2))
    
    for (g1, g2) in year_based_pairs:
        if not remaining_groups[g1] or not remaining_groups[g2]:
            continue
        
        while remaining_groups[g1] and remaining_groups[g2]:
            student1 = remaining_groups[g1].pop(0)
            student2 = remaining_groups[g2].pop(0)
            
            allocation.append({
                'register_number': student1['register_number'],
                'candidate_name': student1['candidate_name'],
                'department': student1['department'],
                'year': student1['year'],
                'phone_number': student1.get('phone_number', ''),
                'room_number': room_no,
                'bench_number': bench_no,
                'seat_position': 'Left'
            })
            
            allocation.append({
                'register_number': student2['register_number'],
                'candidate_name': student2['candidate_name'],
                'department': student2['department'],
                'year': student2['year'],
                'phone_number': student2.get('phone_number', ''),
                'room_number': room_no,
                'bench_number': bench_no,
                'seat_position': 'Right'
            })
            
            bench_no += 1
            if bench_no > 20:
                room_no += 1
                bench_no = 1
    
    # Final remaining (same bench may have same dept if necessary)
    absolute_remaining = []
    for group_list in remaining_groups.values():
        absolute_remaining.extend(group_list)
    
    i = 0
    while i < len(absolute_remaining):
        if i + 1 < len(absolute_remaining):
            student1 = absolute_remaining[i]
            student2 = absolute_remaining[i + 1]
            
            allocation.append({
                'register_number': student1['register_number'],
                'candidate_name': student1['candidate_name'],
                'department': student1['department'],
                'year': student1['year'],
                'phone_number': student1.get('phone_number', ''),
                'room_number': room_no,
                'bench_number': bench_no,
                'seat_position': 'Left'
            })
            
            allocation.append({
                'register_number': student2['register_number'],
                'candidate_name': student2['candidate_name'],
                'department': student2['department'],
                'year': student2['year'],
                'phone_number': student2.get('phone_number', ''),
                'room_number': room_no,
                'bench_number': bench_no,
                'seat_position': 'Right'
            })
            
            bench_no += 1
            if bench_no > 20:
                room_no += 1
                bench_no = 1
            
            i += 2
        else:
            # Odd student - place alone on bench
            student = absolute_remaining[i]
            allocation.append({
                'register_number': student['register_number'],
                'candidate_name': student['candidate_name'],
                'department': student['department'],
                'year': student['year'],
                'phone_number': student.get('phone_number', ''),
                'room_number': room_no,
                'bench_number': bench_no,
                'seat_position': 'Left'
            })
            
            bench_no += 1
            if bench_no > 20:
                room_no += 1
                bench_no = 1
            
            i += 1
    
    return allocation, None

def semester_exam_allocation(students_list):
    """
    Semester Exam Allocation - 1 student per bench
    - 20 benches per classroom = 20 students per classroom
    - EXACTLY 2 departments per room (alternating)
    - Same year per room
    - Sorted by Register Number within each department
    """
    if not students_list:
        return None, "No students to allocate"
    
    # Convert to list of dicts if needed
    students_data = students_list if isinstance(students_list, list) else list(students_list)
    
    # Ensure all students have a Year (default to '1')
    for student in students_data:
        if 'Year' not in student:
            student['Year'] = '1'
    
    # Group by Year
    year_groups = {}
    for student in students_data:
        year = str(student.get('Year', '1'))
        if year not in year_groups:
            year_groups[year] = []
        year_groups[year].append(student)
    
    allocation = []
    current_room = 1
    
    # Process each year separately
    for year_val in sorted(year_groups.keys()):
        year_students = year_groups[year_val]
        
        # Group by department and sort
        dept_dict = {}
        for student in year_students:
            dept = student.get('Department', '')
            if dept not in dept_dict:
                dept_dict[dept] = []
            dept_dict[dept].append(student)
        
        # Sort each department by Register Number
        dept_queues = {}
        for dept in sorted(dept_dict.keys()):
            sorted_students = sorted(dept_dict[dept], 
                                   key=lambda x: str(x.get('Register Number', '')))
            dept_queues[dept] = list(sorted_students)
        
        # Allocate to rooms
        while any(dept_queues.values()):
            # Get available departments with students
            available_depts = [d for d in dept_queues if dept_queues[d]]
            
            if not available_depts:
                break
            
            dept1 = available_depts[0]
            dept2 = available_depts[1] if len(available_depts) > 1 else None
            
            # Fill one room alternating between dept1 and dept2
            bench = 1
            while bench <= 20 and (dept_queues[dept1] or (dept2 and dept_queues[dept2])):
                # Add from dept1
                if dept_queues[dept1] and bench <= 20:
                    student_dict = dept_queues[dept1].pop(0)
                    allocation.append({
                        'register_number': student_dict.get('Register Number', ''),
                        'candidate_name': student_dict.get('Candidate Name', ''),
                        'department': student_dict.get('Department', ''),
                        'year': str(year_val),
                        'phone_number': student_dict.get('Phone Number', ''),
                        'room_number': current_room,
                        'bench_number': bench
                    })
                    bench += 1
                
                # Add from dept2
                if dept2 and dept_queues[dept2] and bench <= 20:
                    student_dict = dept_queues[dept2].pop(0)
                    allocation.append({
                        'register_number': student_dict.get('Register Number', ''),
                        'candidate_name': student_dict.get('Candidate Name', ''),
                        'department': student_dict.get('Department', ''),
                        'year': str(year_val),
                        'phone_number': student_dict.get('Phone Number', ''),
                        'room_number': current_room,
                        'bench_number': bench
                    })
                    bench += 1
            
            current_room += 1
    
    return allocation, None

def perform_seat_allocation(num_classrooms, seats_per_classroom, students_df, exam_type='semester'):
    """
    Main allocation function - routes to exam-type specific logic
    """
    if exam_type == 'internal':
        return internal_exam_allocation(students_df)
    elif exam_type == 'semester':
        return semester_exam_allocation(students_df)
    else:
        # Fallback to internal
        return internal_exam_allocation(students_df)

def perform_seat_allocation_legacy(num_classrooms, seats_per_classroom, students_df):
    """
    Legacy seat allocation algorithm (ordered version)
    - Groups by Department + Year
    - Creates valid pairs where dept1 != dept2 AND year1 != year2
    - Alternates filling from both groups
    - Allocates each student ONLY ONCE
    - Sorted by Register Number for structured, semi-ordered seating
    """
    if students_df is None or students_df.empty:
        return None, "No students to allocate"
    
    # Step 1: Sort by Year and Register Number (not random shuffle)
    df = students_df.sort_values(by=['Year', 'Register Number']).reset_index(drop=True)
    
    # Step 2: Group by Department + Year, maintaining sorted order
    groups = {}
    for (dept, year), group in df.groupby(['Department', 'Year'], sort=False):
        # Sort within group by Register Number to maintain ascending order
        group = group.sort_values('Register Number').reset_index(drop=True)
        group_list = []
        for _, row in group.iterrows():
            group_list.append({
                'register_number': row['Register Number'],
                'candidate_name': row['Candidate Name'],
                'department': row['Department'],
                'year': str(year)
            })
        groups[(dept, year)] = group_list
    
    # Step 3: Create valid pairs (2 dept + 2 different years)
    valid_pairs = []
    for g1, g2 in combinations(groups.keys(), 2):
        dept1, year1 = g1
        dept2, year2 = g2
        
        if dept1 != dept2 and year1 != year2:
            valid_pairs.append((g1, g2))
    
    allocation = []
    allocated_count = 0
    total_students = len(df)
    room_no = 1
    
    # Step 4: Allocate classrooms using valid pairs
    for (g1, g2) in valid_pairs:
        # Check if groups still have students (they might be empty)
        if not groups[g1] or not groups[g2]:
            continue
        
        room_students = []
        
        # Alternate filling - pop from front to maintain ascending order
        while len(room_students) < seats_per_classroom and (groups[g1] or groups[g2]):
            if groups[g1] and len(room_students) < seats_per_classroom:
                student = groups[g1].pop(0)  # Remove from front maintains ascending order
                room_students.append(student)
                allocated_count += 1
            
            if groups[g2] and len(room_students) < seats_per_classroom:
                student = groups[g2].pop(0)  # Remove from front maintains ascending order
                room_students.append(student)
                allocated_count += 1
        
        # Assign seat numbers
        for i, student in enumerate(room_students):
            allocation.append({
                'register_number': student['register_number'],
                'candidate_name': student['candidate_name'],
                'department': student['department'],
                'year': student['year'],
                'room_number': room_no,
                'seat_number': i + 1
            })
        
        room_no += 1
    
    # Step 5: Handle REMAINING students (who don't fit in valid pairs)
    # Re-group remaining students to pair them smartly
    remaining_groups = {}
    for group_key, group_list in groups.items():
        if group_list:  # If this group still has students
            remaining_groups[group_key] = group_list
    
    # Try to create valid pairs from remaining groups (different dept + different year)
    remaining_pairs = []
    for g1, g2 in combinations(remaining_groups.keys(), 2):
        dept1, year1 = g1
        dept2, year2 = g2
        
        if dept1 != dept2 and year1 != year2 and remaining_groups[g1] and remaining_groups[g2]:
            remaining_pairs.append((g1, g2))
    
    # Allocate remaining valid pairs (strict pairing)
    for (g1, g2) in remaining_pairs:
        if not remaining_groups[g1] or not remaining_groups[g2]:
            continue
        
        room_students = []
        
        # Alternate filling from remaining groups
        while len(room_students) < seats_per_classroom and (remaining_groups[g1] or remaining_groups[g2]):
            if remaining_groups[g1] and len(room_students) < seats_per_classroom:
                student = remaining_groups[g1].pop(0)
                room_students.append(student)
            
            if remaining_groups[g2] and len(room_students) < seats_per_classroom:
                student = remaining_groups[g2].pop(0)
                room_students.append(student)
        
        if room_students:
            for i, student in enumerate(room_students):
                allocation.append({
                    'register_number': student['register_number'],
                    'candidate_name': student['candidate_name'],
                    'department': student['department'],
                    'year': student['year'],
                    'room_number': room_no,
                    'seat_number': i + 1
                })
            room_no += 1
    
    # Step 6: After strict pairing, handle any FINAL remaining students
    # Allow pairing by year ONLY (same dept but different years)
    final_remaining_groups = {}
    for group_key, group_list in remaining_groups.items():
        if group_list:
            final_remaining_groups[group_key] = group_list
    
    # Try pairing by year (allow same dept but different year)
    year_based_pairs = []
    for g1, g2 in combinations(final_remaining_groups.keys(), 2):
        dept1, year1 = g1
        dept2, year2 = g2
        
        # Allow if different years (regardless of dept)
        if year1 != year2 and final_remaining_groups[g1] and final_remaining_groups[g2]:
            year_based_pairs.append((g1, g2))
    
    # Allocate year-based pairs
    for (g1, g2) in year_based_pairs:
        if not final_remaining_groups[g1] or not final_remaining_groups[g2]:
            continue
        
        room_students = []
        
        while len(room_students) < seats_per_classroom and (final_remaining_groups[g1] or final_remaining_groups[g2]):
            if final_remaining_groups[g1] and len(room_students) < seats_per_classroom:
                student = final_remaining_groups[g1].pop(0)
                room_students.append(student)
            
            if final_remaining_groups[g2] and len(room_students) < seats_per_classroom:
                student = final_remaining_groups[g2].pop(0)
                room_students.append(student)
        
        if room_students:
            for i, student in enumerate(room_students):
                allocation.append({
                    'register_number': student['register_number'],
                    'candidate_name': student['candidate_name'],
                    'department': student['department'],
                    'year': student['year'],
                    'room_number': room_no,
                    'seat_number': i + 1
                })
            room_no += 1
    
    # Step 7: Any absolutely final remaining (edge case with few students)
    absolute_remaining = []
    for group_list in final_remaining_groups.values():
        absolute_remaining.extend(group_list)
    
    if absolute_remaining:
        i = 0
        while i < len(absolute_remaining):
            room_students = absolute_remaining[i:i+seats_per_classroom]
            
            for j, student in enumerate(room_students):
                allocation.append({
                    'register_number': student['register_number'],
                    'candidate_name': student['candidate_name'],
                    'department': student['department'],
                    'year': student['year'],
                    'room_number': room_no,
                    'seat_number': j + 1
                })
            
            room_no += 1
            i += seats_per_classroom
    
    return allocation, None

def get_department_colors():
    """Return a mapping of departments to colors"""
    colors = {
        'CS': '#FF6B6B',
        'IT': '#4ECDC4',
        'EC': '#45B7D1',
        'ME': '#FFA07A',
        'CE': '#98D8C8',
        'EE': '#F7DC6F',
        'Other': '#BDC3C7'
    }
    return colors

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/admin-login', methods=['GET', 'POST'])
def admin_login():
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '').strip()
        
        # Simple credential check (replace with database validation in production)
        if username == 'admin' and password == 'admin123':
            flash('Admin login successful!', 'success')
            return redirect(url_for('campus_selection'))
        else:
            flash('Invalid username or password!', 'error')
    
    return render_template('admin-login.html')

@app.route('/campus-selection')
def campus_selection():
    return render_template('campus-selection.html')

@app.route('/admin-logout')
def admin_logout():
    """Admin Logout - Clear all admin session data"""
    # Clear admin-related session keys only
    admin_keys = ['exam_config', 'students_count', 'exam_schedule_count', 'students_data', 'allocation_results', 'exam_schedules']
    for key in admin_keys:
        session.pop(key, None)
    session.modified = True
    flash('You have been logged out successfully!', 'success')
    return redirect(url_for('admin_login'))

@app.route('/oncampus-config', methods=['GET', 'POST'])
def oncampus_config():
    """ON-CAMPUS CONFIGURATION PAGE - Step 1"""
    # Enable permanent session for this interaction
    session.permanent = True
    
    if request.method == 'POST':
        try:
            college_name = request.form.get('college_name', '').strip()
            exam_type = request.form.get('exam_type', '').strip()
            exam_date = request.form.get('exam_date', '').strip()
            exam_time = request.form.get('exam_time', '').strip()
            num_classrooms = int(request.form.get('num_classrooms', 0))
            seats_per_classroom = int(request.form.get('seats_per_classroom', 0))
            
            if not college_name or not exam_type or num_classrooms <= 0 or seats_per_classroom <= 0:
                flash('Please enter valid values for all fields!', 'error')
                return render_template('oncampus_config.html')
            
            if exam_type not in ['internal', 'semester']:
                flash('Invalid exam type selected!', 'error')
                return render_template('oncampus_config.html')
            
            # Store in session - CENTRALIZED EXAM CONFIGURATION
            session['exam_config'] = {
                'college_name': college_name,
                'exam_type': exam_type,
                'exam_date': exam_date,
                'exam_time': exam_time,
                'num_classrooms': num_classrooms,
                'seats_per_classroom': seats_per_classroom,
                'total_seats': num_classrooms * seats_per_classroom
            }
            session.modified = True
            
            exam_mode = 'Internal Exam (40 per classroom)' if exam_type == 'internal' else 'Semester Exam (20 per classroom)'
            flash(f'✓ Configuration saved! Mode: {exam_mode} | Exam Date: {exam_date} {exam_time} | Total capacity: {num_classrooms * seats_per_classroom} seats', 'success')
            return redirect(url_for('candidate_upload'))
        except ValueError:
            flash('Please enter valid numbers!', 'error')
    
    config = session.get('exam_config', {})
    return render_template('oncampus_config.html', config=config)

@app.route('/candidate-upload', methods=['GET', 'POST'])
def candidate_upload():
    """CANDIDATE UPLOAD SYSTEM - Step 2"""
    config = session.get('exam_config', {})
    if not config:
        flash('Please configure exam details first!', 'error')
        return redirect(url_for('oncampus_config'))
    
    if request.method == 'POST':
        global students_data
        
        if 'file' not in request.files:
            flash('No file selected!', 'error')
            return render_template('upload.html', config=config)
        
        file = request.files['file']
        if file.filename == '':
            flash('No file selected!', 'error')
            return render_template('upload.html', config=config)
        
        if not allowed_file(file.filename):
            flash('Invalid file type. Please upload .xlsx, .xls, or .csv file', 'error')
            return render_template('upload.html', config=config)
        
        try:
            filename = secure_filename(file.filename)
            filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            file.save(filepath)
            
            # Read file based on extension
            file_ext = filename.rsplit('.', 1)[1].lower()
            if file_ext == 'csv':
                data = read_csv_to_dict(filepath)
            else:  # xlsx or xls
                data = read_excel_to_dict(filepath)
            
            if data is None:
                flash('Error reading file', 'error')
                return render_template('upload.html', config=config)
            
            # Validate and clean data
            cleaned_data, error = validate_student_data(data)
            if error:
                flash(f'Data validation error: {error}', 'error')
                return render_template('upload.html', config=config)
            
            students_data = cleaned_data
            
            # Store in session for persistence across dashboard navigation
            session['students_data'] = cleaned_data
            session['students_count'] = len(cleaned_data)
            session.modified = True
            
            flash(f'✓ Successfully uploaded {len(cleaned_data)} valid candidates!', 'success')
            
            return redirect(url_for('allocate_seats'))
        
        except Exception as e:
            flash(f'Error processing file: {str(e)}', 'error')
            return render_template('upload.html', config=config)
    
    return render_template('upload.html', config=config)

@app.route('/exam-schedule-upload', methods=['GET', 'POST'])
def exam_schedule_upload():
    """EXAM SCHEDULE UPLOAD - Step 2B"""
    # Enable permanent session for this interaction
    session.permanent = True
    
    config = session.get('exam_config', {})
    if not config:
        flash('Please configure exam details first!', 'error')
        return redirect(url_for('oncampus_config'))
    
    if request.method == 'POST':
        global exam_schedules
        
        if 'file' not in request.files:
            flash('No file selected!', 'error')
            return render_template('exam_schedule_upload.html', config=config)
        
        file = request.files['file']
        if file.filename == '':
            flash('No file selected!', 'error')
            return render_template('exam_schedule_upload.html', config=config)
        
        if not allowed_file(file.filename):
            flash('Invalid file type. Please upload .xlsx, .xls, or .csv file', 'error')
            return render_template('exam_schedule_upload.html', config=config)
        
        try:
            filename = secure_filename(file.filename)
            filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            file.save(filepath)
            
            # Read file based on extension
            file_ext = filename.rsplit('.', 1)[1].lower()
            if file_ext == 'csv':
                data = read_csv_to_dict(filepath)
            else:  # xlsx or xls
                data = read_excel_to_dict(filepath)
            
            if data is None:
                flash('Error reading file', 'error')
                return render_template('exam_schedule_upload.html', config=config)
            
            # Validate required columns
            required_columns = ['Year', 'Subject Code', 'Subject Name', 'Exam Date', 'Exam Time']
            if data:
                missing_cols = [col for col in required_columns if col not in data[0].keys()]
                if missing_cols:
                    flash(f'Missing required columns: {", ".join(missing_cols)}', 'error')
                    return render_template('exam_schedule_upload.html', config=config)
            
            # Filter out rows with missing required values
            filtered_data = []
            for row in data:
                if all(row.get(col) is not None and row.get(col) != '' for col in required_columns):
                    filtered_data.append(row)
            exam_schedules = []
            for row in filtered_data:
                schedule_entry = {
                    'year': str(row['Year']).strip(),
                    'department': str(row.get('Department', 'ALL')).strip(),
                    'subject_code': str(row['Subject Code']).strip(),
                    'subject_name': str(row['Subject Name']).strip(),
                    'exam_date': str(row['Exam Date']).strip(),
                    'exam_time': str(row['Exam Time']).strip()
                }
                exam_schedules.append(schedule_entry)
            
            # Store in session for persistence across dashboard navigation
            session['exam_schedules'] = exam_schedules
            session['exam_schedule_count'] = len(exam_schedules)
            session.modified = True
            
            flash(f'✓ Exam schedule uploaded successfully! {len(exam_schedules)} subjects imported', 'success')
            return redirect(url_for('oncampus_dashboard'))
        
        except Exception as e:
            flash(f'Error processing file: {str(e)}', 'error')
            return render_template('exam_schedule_upload.html', config=config)
    
    return render_template('exam_schedule_upload.html', config=config)

@app.route('/allocate-seats')
def allocate_seats():
    """SEAT ALLOCATION ENGINE - Step 3"""
    config = session.get('exam_config', {})
    if not config:
        flash('Please configure exam details first!', 'error')
        return redirect(url_for('oncampus_config'))
    
    # Get students from session (persistent) or global variable
    students_to_allocate = session.get('students_data') or students_data
    
    # Check if students_data is empty
    if not students_to_allocate:
        flash('Please upload candidate data first!', 'error')
        return redirect(url_for('candidate_upload'))
    
    global allocation_results
    
    # Allocate seats based on exam type
    allocation_results, error = perform_seat_allocation(
        config['num_classrooms'],
        config['seats_per_classroom'],
        students_to_allocate,
        config.get('exam_type', 'semester')
    )
    
    if error:
        flash(f'Allocation error: {error}', 'error')
        return redirect(url_for('candidate_upload'))
    
    # Attach subject information to each allocated student
    exam_schedules_from_session = session.get('exam_schedules', [])
    schedules = exam_schedules_from_session if exam_schedules_from_session else exam_schedules
    
    for student in allocation_results:
        student_year = student.get('year', '1')
        student_dept = student.get('department', '')
        
        # Find matching exam schedule
        subject_info = None
        if schedules:
            for schedule in schedules:
                if str(schedule.get('year', '')) == str(student_year):
                    dept = schedule.get('department', 'ALL')
                    if dept == 'ALL' or dept == student_dept:
                        subject_info = schedule
                        break
        
        # Attach subject details
        if subject_info:
            student['subject_code'] = subject_info.get('subject_code', 'N/A')
            student['subject_name'] = subject_info.get('subject_name', 'N/A')
            student['exam_date'] = subject_info.get('exam_date', 'N/A')
        else:
            student['subject_code'] = 'N/A'
            student['subject_name'] = 'N/A'
            student['exam_date'] = 'N/A'
    
    # Store allocation results in session for persistence
    session['allocation_results'] = allocation_results
    session.modified = True
    
    flash(f'✓ Automatic seat allocation completed for {len(allocation_results)} students!', 'success')
    
    return redirect(url_for('view_seating'))

@app.route('/view-seating')
def view_seating():
    """SEATING DISPLAY PAGE - Step 4"""
    config = session.get('exam_config', {})
    
    # Get allocation results from session (persistent) or global variable
    alloc_results = session.get('allocation_results') or allocation_results
    
    if not alloc_results:
        flash('No allocation data available. Please perform seat allocation first!', 'error')
        return redirect(url_for('oncampus_config'))
    
    # Handle search
    search_query = request.args.get('search', '').strip()
    results = alloc_results
    
    if search_query:
        results = [r for r in alloc_results 
                  if str(r['register_number']).lower() == search_query.lower()]
    
    return render_template('seating.html', 
                         results=results, 
                         search_query=search_query,
                         config=config,
                         total_count=len(alloc_results))

@app.route('/classroom-grid')
def classroom_grid():
    """CLASSROOM GRID VISUALIZATION - Professional Bench Layout"""
    config = session.get('exam_config', {})
    
    # Get allocation results from session (persistent) or global variable
    alloc_results = session.get('allocation_results') or allocation_results
    
    if not alloc_results:
        flash('No allocation data available!', 'error')
        return redirect(url_for('oncampus_config'))
    
    # Get search filter if provided
    search_query = request.args.get('search', '').strip()
    results = alloc_results
    
    if search_query:
        results = [r for r in alloc_results 
                  if str(r.get('register_number', '')).lower() == search_query.lower() or
                     str(r.get('candidate_name', '')).lower().find(search_query.lower()) != -1]
    
    # ===== ORGANIZE DATA INTO BENCHES =====
    # Group by room first
    rooms_dict = {}
    for student in results:
        room = student.get('room_number', 1)
        if room not in rooms_dict:
            rooms_dict[room] = []
        rooms_dict[room].append(student)
    
    # For each room, organize students into benches
    rooms_with_benches = {}
    for room_num in sorted(rooms_dict.keys()):
        room_students = rooms_dict[room_num]
        
        # Group students by bench number and add seat numbers
        benches_dict = {}
        for idx, student in enumerate(room_students):
            # Use bench_number if available, otherwise calculate it
            bench = student.get('bench_number')
            if bench is None:
                # Calculate bench: (seat - 1) // 2 + 1
                seat = idx + 1  # Seat within this room (1-40)
                bench = (seat - 1) // 2 + 1
            else:
                bench = int(bench)
            
            # Determine position (left=odd seat, right=even seat)
            if bench not in benches_dict:
                benches_dict[bench] = {'left': None, 'right': None}
            
            # Calculate which seat position this is within the bench
            seat_position = student.get('seat_position', 'left').lower()
            if seat_position == 'right' or seat_position == 'seat_2':
                benches_dict[bench]['right'] = student
            else:
                if benches_dict[bench]['left'] is None:
                    benches_dict[bench]['left'] = student
                else:
                    benches_dict[bench]['right'] = student
        
        # Create all 20 benches for this room (even if some are empty)
        benches = []
        for bench_num in range(1, 21):  # Always create benches 1-20
            bench_obj = {
                'bench_number': bench_num,
                'left_student': benches_dict.get(bench_num, {}).get('left'),
                'right_student': benches_dict.get(bench_num, {}).get('right')
            }
            benches.append(bench_obj)
        
        # Calculate room statistics
        all_depts = set()
        all_years = set()
        all_subjects = set()
        total_students = 0
        
        for bench in benches:
            if bench['left_student']:
                total_students += 1
                all_depts.add(bench['left_student'].get('department', 'N/A'))
                all_years.add(bench['left_student'].get('year', 'N/A'))
                subject_name = bench['left_student'].get('subject_name', 'N/A')
                if subject_name != 'N/A':
                    all_subjects.add(subject_name)
            if bench['right_student']:
                total_students += 1
                all_depts.add(bench['right_student'].get('department', 'N/A'))
                all_years.add(bench['right_student'].get('year', 'N/A'))
                subject_name = bench['right_student'].get('subject_name', 'N/A')
                if subject_name != 'N/A':
                    all_subjects.add(subject_name)
        
        # Build room object with benches and metadata
        rooms_with_benches[room_num] = {
            'room_number': room_num,
            'students_count': total_students,
            'empty_count': 40 - total_students,
            'departments': sorted(list(all_depts)),
            'years': sorted(list(all_years)),
            'subjects': sorted(list(all_subjects)),
            'benches': benches
        }
    
    # Render using new professional bench layout template
    return render_template('classroom_view.html', 
                         rooms_with_benches=rooms_with_benches,
                         search_query=search_query,
                         config=config)

@app.route('/search-student-api')
def search_student_api():
    """API endpoint for student search"""
    reg_num = request.args.get('reg_num', '').strip()
    
    # Get allocation results from session or global
    alloc_results = session.get('allocation_results') or allocation_results
    
    for res in alloc_results:
        if str(res['register_number']).lower() == reg_num.lower():
            return jsonify(res)
    
    return jsonify({'error': 'Student not found'}), 404

@app.route('/export-seating')
def export_seating():
    """Export seating arrangement to Excel - separate sheet for each room"""
    if not allocation_results:
        flash('No data to export!', 'error')
        return redirect(url_for('view_seating'))
    
    try:
        from openpyxl.utils import get_column_letter
        
        filename = f"seating_arrangement_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        
        # Group allocation results by room
        rooms_data = {}
        for result in allocation_results:
            room_no = result.get('room_number', 'Unknown')
            if room_no not in rooms_data:
                rooms_data[room_no] = []
            rooms_data[room_no].append(result)
        
        # Create workbook
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet
        
        # Create summary sheet
        ws_summary = wb.create_sheet('Summary', 0)
        ws_summary.append(['Room Number', 'Total Seats', 'Departments', 'Years'])
        
        header_fill = PatternFill(start_color='4F46E5', end_color='4F46E5', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True, size=11)
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                            top=Side(style='thin'), bottom=Side(style='thin'))
        
        # Format summary header
        for cell in ws_summary[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.border = thin_border
        
        # Add summary data
        for room_no in sorted(rooms_data.keys()):
            room_students = rooms_data[room_no]
            depts = ', '.join(sorted(set(str(s.get('department', 'N/A')) for s in room_students)))
            years = ', '.join(sorted(set(str(s.get('year', 'N/A')) for s in room_students)))
            ws_summary.append([room_no, len(room_students), depts, years])
            
            for cell in ws_summary[ws_summary.max_row]:
                cell.border = thin_border
        
        # Create sheets for each room
        for room_no in sorted(rooms_data.keys()):
            room_students = rooms_data[room_no]
            
            # Determine columns to display
            columns = ['Register Number', 'Candidate Name', 'Department', 'Year']
            
            # Check what fields are available
            if room_students and 'subject_code' in room_students[0]:
                columns.append('Subject Code')
            if room_students and 'subject_name' in room_students[0]:
                columns.append('Subject Name')
            if room_students and 'bench_number' in room_students[0]:
                columns.append('Bench')
            if room_students and 'seat_position' in room_students[0]:
                columns.append('Position')
            
            # Create sheet for room
            ws_room = wb.create_sheet(f'Room {room_no}')
            ws_room.append(columns)
            
            # Format header
            for cell in ws_room[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.border = thin_border
            
            # Map student data to columns
            for student in room_students:
                row_data = [
                    student.get('register_number', ''),
                    student.get('candidate_name', ''),
                    student.get('department', ''),
                    student.get('year', '')
                ]
                
                if 'Subject Code' in columns:
                    row_data.append(student.get('subject_code', 'N/A'))
                if 'Subject Name' in columns:
                    row_data.append(student.get('subject_name', 'N/A'))
                if 'Bench' in columns:
                    row_data.append(student.get('bench_number', ''))
                if 'Position' in columns:
                    row_data.append(student.get('seat_position', ''))
                
                ws_room.append(row_data)
            
            # Format data rows and set column widths
            header_count = len(columns)
            for row_idx, row in enumerate(ws_room.iter_rows(min_row=1, max_row=ws_room.max_row), 1):
                for cell in row:
                    cell.border = thin_border
                    if row_idx > 1:
                        cell.alignment = Alignment(horizontal='left', vertical='center')
            
            # Auto-fit column widths
            for col_idx, column in enumerate(ws_room.columns, 1):
                max_length = len(str(columns[col_idx - 1])) if col_idx <= len(columns) else 0
                for cell in column:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                ws_room.column_dimensions[get_column_letter(col_idx)].width = min(max_length + 2, 50)
            
            # Freeze header
            ws_room.freeze_panes = 'A2'
        
        # Save workbook
        wb.save(filepath)
        
        flash(f'✓ Exported to {filename} ({len(rooms_data)} rooms in separate sheets)', 'success')
        return redirect(url_for('view_seating'))
    except Exception as e:
        import traceback
        print(traceback.format_exc())
        flash(f'Error exporting file: {str(e)}', 'error')
        return redirect(url_for('view_seating'))

@app.route('/export-pdf')
def export_pdf():
    """Export seating arrangement to PDF without student names"""
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from io import BytesIO
    
    if not allocation_results:
        flash('No data to export!', 'error')
        return redirect(url_for('view_seating'))
    
    try:
        # Create PDF in memory
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), 
                               leftMargin=0.5*inch, rightMargin=0.5*inch,
                               topMargin=0.5*inch, bottomMargin=0.5*inch)
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            textColor=colors.HexColor('#667eea'),
            spaceAfter=6,
            alignment=1  # center alignment
        )
        room_style = ParagraphStyle(
            'RoomTitle',
            parent=styles['Heading3'],
            fontSize=12,
            textColor=colors.HexColor('#667eea'),
            spaceAfter=4,
            alignment=0  # left alignment
        )
        
        elements = []
        
        # Title
        config = session.get('exam_config', {})
        title = Paragraph(
            f"<b>Seating Arrangement Report</b>",
            title_style
        )
        elements.append(title)
        college = Paragraph(
            f"<b>{config.get('college_name', 'College')}</b>",
            room_style
        )
        elements.append(college)
        elements.append(Spacer(1, 0.15*inch))
        
        # Group by room
        rooms = {}
        for alloc in allocation_results:
            room = alloc['room_number']
            if room not in rooms:
                rooms[room] = []
            rooms[room].append(alloc)
        
        # Create table for each room
        for room_no in sorted(rooms.keys()):
            room_data = rooms[room_no]
            
            # Room header
            room_title = Paragraph(f"<b>Room {room_no}</b>", room_style)
            elements.append(room_title)
            elements.append(Spacer(1, 0.08*inch))
            
            # Create table data
            table_data = [['Bench/Seat #', 'Register Number', 'Department', 'Year', 'Subject Code']]
            
            # Sort by bench_number (safe access with fallback)
            for seat in sorted(room_data, key=lambda x: x.get('bench_number', x.get('bench_no', 0))):
                # Build bench/seat display text safely
                bench_num = seat.get('bench_number', seat.get('bench_no', 'N/A'))
                seat_pos = seat.get('seat_position', '')
                
                if seat_pos and seat_pos.lower() in ['left', 'right']:
                    # Internal exam: show bench and seat position
                    bench_display = f"Bench {bench_num} ({seat_pos})"
                else:
                    # Semester exam: just show bench number
                    bench_display = f"Bench {bench_num}"
                
                table_data.append([
                    bench_display,
                    str(seat['register_number']),
                    str(seat['department']),
                    str(seat['year']),
                    str(seat.get('subject_code', 'N/A'))
                ])
            
            # Create table with better column widths
            col_widths = [1.2*inch, 1.6*inch, 1.3*inch, 0.8*inch, 1.2*inch]
            table = Table(table_data, colWidths=col_widths)
            
            table.setStyle(TableStyle([
                # Header styling
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#667eea')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                ('VALIGN', (0, 0), (-1, 0), 'MIDDLE'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('TOPPADDING', (0, 0), (-1, 0), 12),
                
                # Data styling
                ('ALIGN', (0, 1), (0, -1), 'CENTER'),  # First column center
                ('ALIGN', (1, 1), (-1, -1), 'CENTER'),  # Rest columns center
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 9),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f0f4ff')]),
                
                # Grid styling
                ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#cccccc')),
                ('TOPPADDING', (0, 1), (-1, -1), 10),
                ('BOTTOMPADDING', (0, 1), (-1, -1), 10),
                ('LEFTPADDING', (0, 0), (-1, -1), 8),
                ('RIGHTPADDING', (0, 0), (-1, -1), 8),
            ]))
            
            elements.append(table)
            
            # Add page break after each room except the last one
            if room_no != sorted(rooms.keys())[-1]:
                elements.append(PageBreak())
        
        # Build PDF
        doc.build(elements)
        
        # Return as download
        buffer.seek(0)
        filename = f"seating_arrangement_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        return send_file(
            buffer,
            mimetype='application/pdf',
            as_attachment=True,
            download_name=filename
        )
    except Exception as e:
        flash(f'Error generating PDF: {str(e)}', 'error')
        return redirect(url_for('view_seating'))

# ========== STUDENT PORTAL ROUTES ==========

@app.route('/student-login', methods=['GET', 'POST'])
def student_login():
    """Student Login Page - Authenticate using Register Number and Year/DOB"""
    if request.method == 'POST':
        reg_num = request.form.get('register_number', '').strip().upper()
        year_or_dob = request.form.get('year_or_dob', '').strip()
        
        # Search for student in allocation_results
        student_found = None
        for alloc in allocation_results:
            if str(alloc.get('register_number', '')).upper() == reg_num:
                # Additional validation: check year (simple format "1", "2", "3", "4")
                if year_or_dob == str(alloc.get('year', '')):
                    student_found = alloc
                    break
        
        if student_found:
            # Store in session
            session['student_logged_in'] = True
            session['student_register_number'] = reg_num
            session['student_name'] = student_found.get('candidate_name', 'Student')
            session['student_year'] = student_found.get('year', '')
            session.modified = True
            
            flash(f'Welcome {student_found.get("candidate_name", "Student")}!', 'success')
            return redirect(url_for('student_dashboard'))
        else:
            flash('Invalid Register Number or Year. Please try again.', 'error')
    
    return render_template('student-login.html')

@app.route('/student-dashboard')
def student_dashboard():
    """Student Dashboard - Display seating details and exam schedule"""
    if not session.get('student_logged_in'):
        flash('Please login first!', 'error')
        return redirect(url_for('student_login'))
    
    reg_num = session.get('student_register_number', '').upper()
    
    # Find student's allocation data
    student_data = None
    bench_mate = None
    
    for alloc in allocation_results:
        if str(alloc.get('register_number', '')).upper() == reg_num:
            student_data = alloc
            break
    
    if not student_data:
        flash('Seating allocation not found!', 'error')
        return redirect(url_for('student_login'))
    
    # Find bench mate for internal exam (same bench, Opposite position)
    if 'seat_position' in student_data:  # Internal exam
        for alloc in allocation_results:
            if (str(alloc.get('register_number', '')).upper() != reg_num and
                alloc.get('room_number') == student_data.get('room_number') and
                alloc.get('bench_number') == student_data.get('bench_number')):
                bench_mate = alloc
                break
    
    # Get centralized config from session
    config = session.get('exam_config', {})
    
    # Get exam schedule for student (FEATURE 4: Automatic subject matching)
    exam_schedule = get_exam_schedule_for_student(
        student_data.get('year'),
        student_data.get('department')
    )
    
    # If exam_date and exam_time are not set in config, derive from exam_schedule
    if exam_schedule:
        if not config.get('exam_date') or config.get('exam_date') == '':
            config['exam_date'] = exam_schedule.get('exam_date', 'Not Set')
        if not config.get('exam_time') or config.get('exam_time') == '':
            config['exam_time'] = exam_schedule.get('exam_time', 'Not Set')
    
    return render_template('student-dashboard.html', 
                         student=student_data,
                         bench_mate=bench_mate,
                         config=config,
                         exam_schedule=exam_schedule)  # New: Pass exam schedule

@app.route('/student-logout')
def student_logout():
    """Student Logout"""
    session.clear()
    flash('You have been logged out!', 'info')
    return redirect(url_for('student_login'))

@app.route('/student-download-slip')
def student_download_slip():
    """Download Seating Slip as PDF"""
    if not session.get('student_logged_in'):
        return redirect(url_for('student_login'))
    
    reg_num = session.get('student_register_number', '').upper()
    
    # Find student data
    student_data = None
    for alloc in allocation_results:
        if str(alloc.get('register_number', '')).upper() == reg_num:
            student_data = alloc
            break
    
    if not student_data:
        flash('Seating data not found!', 'error')
        return redirect(url_for('student_dashboard'))
    
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.lib.units import inch, cm
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image as RLImage
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.enums import TA_CENTER, TA_LEFT
        from io import BytesIO
        
        # Create PDF in memory
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4,
                               leftMargin=0.8*inch, rightMargin=0.8*inch,
                               topMargin=0.8*inch, bottomMargin=0.8*inch)
        
        styles = getSampleStyleSheet()
        elements = []
        
        # Get config
        config = session.get('exam_config', {})
        college_name = config.get('college_name', 'College')
        exam_type = session.get('exam_config', {}).get('exam_type', 'Semester Exam')
        exam_type_display = 'Internal Exam' if exam_type == 'internal' else 'Semester Exam'
        
        # Title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            textColor=colors.HexColor('#1a1a2e'),
            spaceAfter=6,
            alignment=TA_CENTER,
            fontName='Helvetica-Bold'
        )
        
        college_title = Paragraph(college_name, title_style)
        elements.append(college_title)
        
        # Subtitle
        subtitle_style = ParagraphStyle(
            'CustomSubtitle',
            parent=styles['Heading3'],
            fontSize=14,
            textColor=colors.HexColor('#2ebf91'),
            spaceAfter=20,
            alignment=TA_CENTER,
            fontName='Helvetica-Bold'
        )
        
        subtitle = Paragraph('SEAT ALLOCATION SLIP', subtitle_style)
        elements.append(subtitle)
        
        # QR Code (bonus)
        try:
            qr_data = f"REG:{reg_num}|ROOM:{student_data.get('room_number')}|BENCH:{student_data.get('bench_number')}"
            qr = qrcode.QRCode(version=1, box_size=5, border=2)
            qr.add_data(qr_data)
            qr.make(fit=True)
            qr_img = qr.make_image(fill_color="black", back_color="white")
            
            # Save QR to buffer
            qr_buffer = BytesIO()
            qr_img.save(qr_buffer, format='PNG')
            qr_buffer.seek(0)
            
            # Add QR code to PDF
            elements.append(Spacer(1, 0.2*inch))
            qr_rl_img = RLImage(qr_buffer, width=1.2*inch, height=1.2*inch)
            elements.append(qr_rl_img)
            elements.append(Spacer(1, 0.2*inch))
        except:
            pass  # If QR fails, continue without it
        
        # Student Details Table
        details_data = [
            ['Field', 'Details'],
            ['Register Number', str(student_data.get('register_number', 'N/A'))],
            ['Student Name', str(student_data.get('candidate_name', 'N/A'))],
            ['Department', str(student_data.get('department', 'N/A'))],
            ['Year', str(student_data.get('year', 'N/A'))],
            ['Exam Type', exam_type_display],
        ]
        
        details_table = Table(details_data, colWidths=[2*inch, 3*inch])
        details_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a1a2e')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 11),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#f8f9fa')),
            ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#dee2e6')),
            ('FONTSIZE', (0, 1), (-1, -1), 10),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8f9fa')]),
            ('TOPPADDING', (0, 1), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 10),
        ]))
        
        elements.append(details_table)
        elements.append(Spacer(1, 0.3*inch))
        
        # Seating Details
        seating_title = Paragraph('SEATING DETAILS', subtitle_style)
        elements.append(seating_title)
        elements.append(Spacer(1, 0.15*inch))
        
        seating_data = [
            ['Attribute', 'Value'],
            ['Room Number', str(student_data.get('room_number', 'N/A'))],
            ['Bench Number', str(student_data.get('bench_number', 'N/A'))],
        ]
        
        # Add seat position for internal exam
        if 'seat_position' in student_data:
            seating_data.append(['Seat Position', str(student_data.get('seat_position', 'N/A'))])
        
        seating_data.append(['Allocated On', datetime.now().strftime('%d-%m-%Y at %H:%M:%S')])
        
        seating_table = Table(seating_data, colWidths=[2*inch, 3*inch])
        seating_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2ebf91')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 11),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#f8f9fa')),
            ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#dee2e6')),
            ('FONTSIZE', (0, 1), (-1, -1), 10),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8f9fa')]),
            ('TOPPADDING', (0, 1), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 10),
        ]))
        
        elements.append(seating_table)
        elements.append(Spacer(1, 0.3*inch))
        
        # Exam Schedule Details (if available)
        exam_schedule = get_exam_schedule_for_student(
            student_data.get('year'),
            student_data.get('department')
        )
        
        if exam_schedule:
            schedule_title = Paragraph('EXAM SCHEDULE', subtitle_style)
            elements.append(schedule_title)
            elements.append(Spacer(1, 0.15*inch))
            
            schedule_data = [
                ['Field', 'Details'],
                ['Subject Code', str(exam_schedule.get('subject_code', 'N/A'))],
                ['Subject Name', str(exam_schedule.get('subject_name', 'N/A'))],
                ['Exam Date', str(exam_schedule.get('exam_date', 'N/A'))],
                ['Exam Time', str(exam_schedule.get('exam_time', 'N/A'))],
            ]
            
            schedule_table = Table(schedule_data, colWidths=[2*inch, 3*inch])
            schedule_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3498db')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 11),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#f8f9fa')),
                ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#dee2e6')),
                ('FONTSIZE', (0, 1), (-1, -1), 10),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8f9fa')]),
                ('TOPPADDING', (0, 1), (-1, -1), 10),
                ('BOTTOMPADDING', (0, 1), (-1, -1), 10),
            ]))
            
            elements.append(schedule_table)
            elements.append(Spacer(1, 0.3*inch))
        
        # Footer
        footer_style = ParagraphStyle(
            'FooterStyle',
            parent=styles['Normal'],
            fontSize=9,
            textColor=colors.HexColor('#868e96'),
            alignment=TA_CENTER
        )
        footer = Paragraph('This is an officially generated seating slip. Please preserve it for exam day.<br/>Report to your assigned room and bench before the exam starts.', footer_style)
        elements.append(footer)
        
        # Build PDF
        doc.build(elements)
        
        # Return as download
        buffer.seek(0)
        filename = f"seating_slip_{reg_num}_{datetime.now().strftime('%Y%m%d')}.pdf"
        return send_file(
            buffer,
            mimetype='application/pdf',
            as_attachment=True,
            download_name=filename
        )
    
    except Exception as e:
        import traceback
        print(traceback.format_exc())
        flash(f'Error generating PDF: {str(e)}', 'error')
        return redirect(url_for('student_dashboard'))

@app.route('/dashboard')
def dashboard():
    return render_template('dashboard.html')

@app.route('/oncampus-dashboard')
def oncampus_dashboard():
    # Enable permanent session for this interaction
    session.permanent = True
    
    # Get exam config and data from session (persistent)
    config = session.get('exam_config', {})
    students_from_session = session.get('students_data')
    if students_from_session is None:
        students_from_session = students_data
    
    allocation_from_session = session.get('allocation_results') or allocation_results
    
    stats = {
        'total_students': len(students_from_session),
        'total_classrooms': config.get('num_classrooms', 10),
        'total_seats': config.get('total_seats', 450),
        'allocated_seats': len(allocation_from_session)
    }
    
    # Calculate session status indicators
    # Use actual data presence as source of truth
    students_count = len(students_from_session)
    students_uploaded = students_count > 0
    
    session_status = {
        'config_completed': bool(config),
        'students_uploaded': students_uploaded,
        'students_count': students_count,
        'exam_schedule_uploaded': len(session.get('exam_schedules', [])) > 0,
        'exam_schedule_count': len(session.get('exam_schedules', [])),
        'allocation_completed': len(allocation_from_session) > 0,
        'allocation_count': len(allocation_from_session)
    }
    
    return render_template('oncampus_dashboard.html', 
                         stats=stats, 
                         allocation_results=allocation_from_session,
                         config=config,
                         session_status=session_status)

@app.route('/offcampus-dashboard')
def offcampus_dashboard():
    stats = {
        'total_students': len(students_data),
        'total_venues': 5,         # Placeholder
        'total_seats': 300,        # Placeholder
        'allocated_seats': len(allocation_results)
    }
    return render_template('offcampus_dashboard.html', stats=stats)

@app.route('/upload-students', methods=['POST'])
def upload_students():
    global students_data
    # Enable permanent session for this interaction
    session.permanent = True
    
    if 'file' not in request.files:
        flash('No file part', 'error')
        return redirect(url_for('oncampus_dashboard'))
    
    file = request.files['file']
    if file.filename == '':
        flash('No selected file', 'error')
        return redirect(url_for('oncampus_dashboard'))
    
    if file and allowed_file(file.filename):
        filename = secure_filename(file.filename)
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        file.save(filepath)
        
        try:
            # Get file extension
            file_ext = filename.rsplit('.', 1)[1].lower()
            
            if file_ext == 'csv':
                data = read_csv_to_dict(filepath)
            elif file_ext in ['xlsx', 'xls']:
                data = read_excel_to_dict(filepath)
            else:
                flash(f'Unsupported file format: {file_ext}', 'error')
                return redirect(url_for('oncampus_dashboard'))
            
            if data is None:
                flash('Error reading file', 'error')
                return redirect(url_for('oncampus_dashboard'))
            
            # Required columns: Register Number, (Name or Student Name), Department
            # Accept both "Name" and "Student Name" for student name column
            required_columns = ['Register Number', 'Department']
            if data:
                available_cols = list(data[0].keys())
                missing_columns = [col for col in required_columns if col not in available_cols]
                
                # Check for at least one of the name column variants
                has_name_column = 'Name' in available_cols or 'Student Name' in available_cols
                if not has_name_column:
                    missing_columns.append("Name or Student Name")
                
                if missing_columns:
                    flash(f"Missing required columns: {', '.join(missing_columns)}. Your file has: {', '.join(available_cols)}", 'error')
                    return redirect(url_for('oncampus_dashboard'))
                
                # Normalize column names if needed (rename "Name" to "Student Name" for consistency)
                if 'Name' in available_cols and 'Student Name' not in available_cols:
                    for row in data:
                        if 'Name' in row:
                            row['Student Name'] = row.pop('Name')
            
            students_data = data
            
            # Store in session for persistence across dashboard navigation
            session['students_data'] = data
            session.modified = True
            
            flash(f'✓ Successfully uploaded {len(students_data)} students.', 'success')
        except Exception as e:
            flash(f'Error processing file: {str(e)}', 'error')
            
        return redirect(url_for('oncampus_dashboard'))
    
    flash('Invalid file type. Please upload CSV or Excel (.xlsx, .xls)', 'error')
    return redirect(url_for('oncampus_dashboard'))

@app.route('/generate-seating')
def generate_seating():
    global students_data, allocation_results
    
    # Get students from session (persistent) or global variable
    students_to_allocate = session.get('students_data') or students_data
    
    if not students_to_allocate:
        flash('No student data available to allocate.', 'error')
        return redirect(url_for('oncampus_dashboard'))
    
    allocation_results = []
    classrooms = ['Hall A', 'Hall B', 'Hall C', 'Hall D']
    benches_per_classroom = 25
    seats_per_bench = 2
    
    # Group students by department
    departments = {}
    for student in students_to_allocate:
        dept = student.get('Department', 'Unknown')
        if dept not in departments:
            departments[dept] = []
        departments[dept].append(student)
    
    # Get list of departments
    dept_list = list(departments.keys())
    
    if len(dept_list) < 2:
        flash('Need at least 2 departments for seating arrangement.', 'error')
        return redirect(url_for('oncampus_dashboard'))
    
    classroom_index = 0
    bench_index = 1
    
    # Pair departments and allocate to classrooms
    dept_pairs = []
    for i in range(0, len(dept_list), 2):
        if i + 1 < len(dept_list):
            dept_pairs.append((dept_list[i], dept_list[i + 1]))
        else:
            # If odd number of departments, pair last with first
            dept_pairs.append((dept_list[i], dept_list[0]))
    
    # Allocate seating for each department pair
    for dept1, dept2 in dept_pairs:
        if classroom_index >= len(classrooms):
            break
        
        students_dept1 = departments[dept1][:]
        students_dept2 = departments[dept2][:]
        
        # Allocate benches (25 benches per classroom)
        for bench_num in range(benches_per_classroom):
            if classroom_index >= len(classrooms):
                break
            
            current_classroom = classrooms[classroom_index]
            
            # Try to get one student from each department
            student1 = None
            student2 = None
            
            if students_dept1:
                student1 = students_dept1.pop(0)
            if students_dept2:
                student2 = students_dept2.pop(0)
            
            # If one department runs out, move to next classroom
            if not student1 or not student2:
                classroom_index += 1
                if classroom_index >= len(classrooms):
                    break
                current_classroom = classrooms[classroom_index]
                
                # Allocate remaining students to new classroom
                if student1:
                    allocation_results.append({
                        'register_number': student1.get('Register Number'),
                        'student_name': student1.get('Student Name'),
                        'department': student1.get('Department'),
                        'classroom': current_classroom,
                        'bench_number': f"B-{bench_num + 1}",
                        'seat_number': 'S-1'
                    })
                if student2:
                    allocation_results.append({
                        'register_number': student2.get('Register Number'),
                        'student_name': student2.get('Student Name'),
                        'department': student2.get('Department'),
                        'classroom': current_classroom,
                        'bench_number': f"B-{bench_num + 1}",
                        'seat_number': 'S-2'
                    })
                continue
            
            # Allocate both students to same bench (different departments)
            # Seat 1 - Student from Dept 1
            allocation_results.append({
                'register_number': student1.get('Register Number'),
                'student_name': student1.get('Student Name'),
                'department': student1.get('Department'),
                'classroom': current_classroom,
                'bench_number': f"B-{bench_num + 1}",
                'seat_number': 'S-1'
            })
            
            # Seat 2 - Student from Dept 2
            allocation_results.append({
                'register_number': student2.get('Register Number'),
                'student_name': student2.get('Student Name'),
                'department': student2.get('Department'),
                'classroom': current_classroom,
                'bench_number': f"B-{bench_num + 1}",
                'seat_number': 'S-2'
            })
        
        classroom_index += 1
    
    if allocation_results:
        # Store allocation results in session for persistence
        session['allocation_results'] = allocation_results
        session.modified = True
        
        flash(f'✓ Successfully generated smart seating for {len(allocation_results)} students! ({len(dept_pairs)} department pairs)', 'success')
    else:
        flash('Could not generate seating arrangement.', 'error')
    
    return redirect(url_for('view_results'))

@app.route('/view-results')
def view_results():
    # Get allocation results from session or global
    alloc_results = session.get('allocation_results') or allocation_results
    
    search_query = request.args.get('search', '').strip()
    results = alloc_results
    
    if search_query:
        results = [r for r in alloc_results if str(r['register_number']) == search_query]
        
    return render_template('view_results.html', results=results, search_query=search_query)

@app.route('/classroom-visualization')
def classroom_visualization():
    """Display results in realistic classroom layout"""
    # Get allocation results from session or global
    alloc_results = session.get('allocation_results') or allocation_results
    
    if not alloc_results:
        flash('No allocation data available!', 'error')
        return redirect(url_for('oncampus_dashboard'))
    
    search_query = request.args.get('search', '').strip()
    results = alloc_results
    
    if search_query:
        results = [r for r in alloc_results 
                  if str(r.get('register_number', '')).lower() == search_query.lower() or
                     str(r.get('student_name', '')).lower().find(search_query.lower()) != -1]
    
    return render_template('classroom_view.html', results=results, search_query=search_query)

@app.route('/search-student')
def search_student():
    # Get allocation results from session or global
    alloc_results = session.get('allocation_results') or allocation_results
    
    reg_num = request.args.get('reg_num', '').strip()
    for res in alloc_results:
        if str(res['register_number']) == reg_num:
            return jsonify(res)
    return jsonify({'error': 'Student not found'}), 404

@app.route('/student-allotment')
def student_allotment():
    return render_template('student-allotment.html')

@app.route('/send-sms-notifications', methods=['POST'])
def send_sms_notifications_route():
    """Send SMS notifications to all allocated students"""
    try:
        config = session.get('exam_config', {})
        
        # Get allocation results from session or global
        alloc_results = session.get('allocation_results') or allocation_results
        
        if not alloc_results:
            logger.warning("SMS request: No allocations found")
            return jsonify({
                "status": "error",
                "success_count": 0,
                "failed_count": 0,
                "skipped_count": 0,
                "total": 0,
                "message": "No students allocated yet. Please allocate seats first.",
                "verified_count": len(VERIFIED_TEST_NUMBERS)
            }), 400
        
        logger.info("\n📱 SMS NOTIFICATIONS ENDPOINT CALLED")
        
        # Send SMS notifications with robust error handling
        result = send_bulk_sms_notifications()
        
        return jsonify(result), 200
    
    except Exception as e:
        logger.error(f"❌ Unexpected error: {str(e)}")
        return jsonify({
            "status": "error",
            "success_count": 0,
            "failed_count": 0,
            "skipped_count": 0,
            "total": 0,
            "message": f"Server error: {str(e)}",
            "verified_count": len(VERIFIED_TEST_NUMBERS)
        }), 500

if __name__ == '__main__':
    # Get PORT from environment variable (set by Render) or default to 5000
    port = int(os.environ.get('PORT', 5000))
    # Disable debug in production (Render) - this will be False on live deployment
    debug = os.environ.get('FLASK_ENV', 'development') == 'development'
    app.run(debug=debug, host='0.0.0.0', port=port)
