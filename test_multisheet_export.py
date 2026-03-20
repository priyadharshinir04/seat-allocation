"""
Test script to verify Excel export with separate sheets for each room
"""

import pandas as pd
import os
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Create test allocation data (3 rooms with 10 seats each)
allocation_results = []

# Room 1 - Semester exam (1 student per bench)
for i in range(1, 11):
    dept = 'CSE' if i % 2 == 1 else 'IT'
    allocation_results.append({
        'register_number': f'21CS{1000+i}' if dept == 'CSE' else f'21IT{1000+i}',
        'candidate_name': f'Student {i}',
        'department': dept,
        'year': '1',
        'room_number': 1,
        'bench_number': i
    })

# Room 2 - Semester exam (1 student per bench)
for i in range(1, 11):
    dept = 'CSE' if i % 2 == 1 else 'IT'
    allocation_results.append({
        'register_number': f'21CS{2000+i}' if dept == 'CSE' else f'21IT{2000+i}',
        'candidate_name': f'Student {i+10}',
        'department': dept,
        'year': '1',
        'room_number': 2,
        'bench_number': i
    })

# Room 3 - Semester exam (1 student per bench)
for i in range(1, 11):
    dept = 'CSE' if i % 2 == 1 else 'IT'
    allocation_results.append({
        'register_number': f'21CS{3000+i}' if dept == 'CSE' else f'21IT{3000+i}',
        'candidate_name': f'Student {i+20}',
        'department': dept,
        'year': '1',
        'room_number': 3,
        'bench_number': i
    })

print("=" * 80)
print("TESTING MULTI-SHEET EXCEL EXPORT")
print("=" * 80)
print(f"\nTest Data Created:")
print(f"  - 3 Rooms")
print(f"  - 10 seats per room")
print(f"  - Total: {len(allocation_results)} students")
print(f"\nSample allocation:")
for i in range(3):
    alloc = allocation_results[i]
    print(f"  Room {alloc['room_number']}, Bench {alloc['bench_number']}: {alloc['register_number']} ({alloc['department']})")

# Group by room
rooms_data = {}
for result in allocation_results:
    room_no = result.get('room_number', 'Unknown')
    if room_no not in rooms_data:
        rooms_data[room_no] = []
    rooms_data[room_no].append(result)

print(f"\n" + "=" * 80)
print("CREATING MULTI-SHEET EXCEL FILE")
print("=" * 80)

filename = f"test_seating_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
filepath = os.path.join('.', filename)

try:
    with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
        # Create summary sheet
        summary_data = []
        for room_no in sorted(rooms_data.keys()):
            room_students = rooms_data[room_no]
            summary_data.append({
                'Room Number': room_no,
                'Total Seats': len(room_students),
                'Departments': ', '.join(sorted(set(s['department'] for s in room_students))),
                'Years': ', '.join(sorted(set(str(s['year']) for s in room_students)))
            })
        
        summary_df = pd.DataFrame(summary_data)
        summary_df.to_excel(writer, sheet_name='Summary', index=False)
        print(f"\n✓ Created 'Summary' sheet")
        
        # Create a sheet for each room
        for room_no in sorted(rooms_data.keys()):
            room_students = rooms_data[room_no]
            df_room = pd.DataFrame(room_students)
            
            # Select columns
            columns_to_show = ['register_number', 'candidate_name', 'department', 'year', 'bench_number']
            df_room_display = df_room[columns_to_show].copy()
            
            # Rename columns
            rename_map = {
                'register_number': 'Register Number',
                'candidate_name': 'Candidate Name',
                'department': 'Department',
                'year': 'Year',
                'bench_number': 'Bench'
            }
            df_room_display = df_room_display.rename(columns=rename_map)
            df_room_display = df_room_display.sort_values('Bench')
            
            # Write to sheet
            sheet_name = f'Room {room_no}'
            df_room_display.to_excel(writer, sheet_name=sheet_name, index=False)
            
            # Format the sheet
            worksheet = writer.sheets[sheet_name]
            
            # Header styling
            header_fill = PatternFill(start_color='4F46E5', end_color='4F46E5', fill_type='solid')
            header_font = Font(color='FFFFFF', bold=True, size=11)
            header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            # Border style
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Apply header formatting
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
                cell.border = thin_border
            
            # Apply data formatting
            for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
                for cell in row:
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='left', vertical='center')
            
            # Auto-fit columns
            for column in worksheet.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
            # Freeze header
            worksheet.freeze_panes = 'A2'
            
            print(f"✓ Created 'Room {room_no}' sheet with {len(room_students)} students")
    
    print(f"\n" + "=" * 80)
    print("VERIFICATION")
    print("=" * 80)
    
    # Load and verify the file
    wb = load_workbook(filepath)
    print(f"\n✓ File created: {filename}")
    print(f"✓ Total sheets: {len(wb.sheetnames)}")
    print(f"✓ Sheet names: {', '.join(wb.sheetnames)}")
    
    # Display sheet contents
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"\n  {sheet_name}:")
        print(f"    - Headers: {[cell.value for cell in ws[1]]}")
        print(f"    - Data rows: {ws.max_row - 1}")
        
        # Show first 3 rows
        for i, row in enumerate(ws.iter_rows(min_row=2, max_row=4, values_only=True), 1):
            print(f"    - Row {i}: {row}")
    
    print(f"\n" + "=" * 80)
    print("✓ MULTI-SHEET EXPORT TEST SUCCESSFUL!")
    print("=" * 80)
    print(f"\nFile location: {os.path.abspath(filepath)}")
    print("\nFeatures:")
    print("  ✓ Summary sheet with all rooms overview")
    print("  ✓ Separate sheet for each room")
    print("  ✓ Professional formatting (colored headers, borders)")
    print("  ✓ Auto-fitted columns")
    print("  ✓ Frozen header row")
    print("  ✓ Sorted by Bench number")
    print("\nYou can now open this file in Excel to verify the layout!")

except Exception as e:
    print(f"\n❌ ERROR: {str(e)}")
    import traceback
    traceback.print_exc()
